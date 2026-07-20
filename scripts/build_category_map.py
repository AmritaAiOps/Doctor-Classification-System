"""Exports the 'Category codes' sheet into config/category_map.json.

Run manually whenever the bundled Category Codes reference sheet changes.
Not run per-request; the app loads the generated JSON at runtime.
"""
import json
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
SOURCE_WORKBOOK = BASE_DIR / "config" / "final_output_template.xlsx"
SHEET_NAME = "Category codes"
OUTPUT_PATH = BASE_DIR / "config" / "category_map.json"

BUCKET_HEADER_ROW = {"CPR", "General", "ECHS", "TPA", "P CARD FUND"}


def normalize(value: str) -> str:
    """Strips ALL non-alphanumeric characters (not just whitespace), matching
    backend.stages.category_mapping.normalize_loose -- otherwise HIS export
    values punctuated slightly differently than this master sheet (e.g.
    "ESI2025" vs "ESI - 2025") fail to match and silently fall through as
    unmatched."""
    return re.sub(r"[^a-z0-9]", "", value.strip().lower())


def build_map():
    wb = openpyxl.load_workbook(SOURCE_WORKBOOK, data_only=True)
    ws = wb[SHEET_NAME]

    entries = {}
    current_region = None
    bucket_columns = None  # {col_index: bucket_name}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [cell.value for cell in row]
        first = (values[0] or "").strip() if isinstance(values[0], str) else values[0]

        if first == "Domestic":
            current_region = "Domestic"
            bucket_columns = None
            continue
        if first == "International":
            current_region = "International"
            bucket_columns = None
            continue

        stripped_values = [v.strip() if isinstance(v, str) else v for v in values]
        if current_region and bucket_columns is None and stripped_values and stripped_values[0] in BUCKET_HEADER_ROW:
            bucket_columns = {i: v for i, v in enumerate(stripped_values) if v}
            continue

        if current_region and bucket_columns:
            for col_idx, bucket_name in bucket_columns.items():
                if col_idx >= len(values):
                    continue
                raw = values[col_idx]
                if not raw or not isinstance(raw, str):
                    continue
                key = normalize(raw)
                if not key:
                    continue
                entries[key] = {"bucket": bucket_name, "region": current_region}

    OUTPUT_PATH.write_text(json.dumps(entries, indent=2, sort_keys=True))
    print(f"Wrote {len(entries)} mappings to {OUTPUT_PATH}")


if __name__ == "__main__":
    build_map()
