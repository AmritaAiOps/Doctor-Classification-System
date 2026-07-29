import openpyxl
import pytest

from backend.stages.final_output import write_final_output, SHEET_NAME


@pytest.fixture
def template_path(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["E4"] = "Bed Strength"
    ws["E5"] = "Beds Occupied"
    ws["E8"] = "Volumes"
    ws["E29"] = "Long Stay Patients"
    ws["E32"] = "OP"
    ws["F8"] = 999  # section-header row, never in ROW_MAP -- must stay untouched
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return path


def test_write_final_output_writes_known_rows(template_path, tmp_path):
    values = {
        "Bed Strength": 1000,
        "Beds Occupied": 966,
        "OP Billing": 9923868.6,  # rounded to a whole number on write
        "Long Stay Patients": 14,
    }
    output_path = tmp_path / "output.xlsx"
    result_path = write_final_output(template_path, output_path, values, date_column="F")

    wb = openpyxl.load_workbook(result_path)
    ws = wb[SHEET_NAME]
    assert ws["F4"].value == 1000
    assert ws["F5"].value == 966
    assert ws["F32"].value == 9923869
    assert ws["F29"].value == 14


def test_write_final_output_leaves_non_data_rows_untouched(template_path, tmp_path):
    values = {"Bed Strength": 1000}
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="F")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    # Row 8 ("Volumes" section header) is not in ROW_MAP -- must be untouched
    assert ws["F8"].value == 999


def test_write_final_output_missing_key_leaves_cell_untouched(template_path, tmp_path):
    values = {}  # nothing supplied
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="F")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    assert ws["F4"].value is None


def test_write_final_output_uses_given_date_column(template_path, tmp_path):
    values = {"Bed Strength": 1234}
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="G")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    assert ws["G4"].value == 1234
    assert ws["F4"].value is None


def test_write_final_output_returns_output_path(template_path, tmp_path):
    output_path = tmp_path / "output.xlsx"
    result = write_final_output(template_path, output_path, {"Bed Strength": 1}, date_column="F")
    assert str(result) == str(output_path)


def test_write_final_output_writes_daily_avg_and_mtd_proj_columns(template_path, tmp_path):
    values = {"Bed Strength": 1000, "Beds Occupied": 1028}
    mtd_columns = {
        "Bed Strength": {"daily_avg": 1000, "mtd_proj": 1000},
        "Beds Occupied": {"daily_avg": 964.6, "mtd_proj": 964.6},
    }
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="F", mtd_columns=mtd_columns)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    # date_column="F" -> Daily Average in G, MTD (Proj) in H
    assert ws["G4"].value == 1000
    assert ws["H4"].value == 1000
    assert ws["G5"].value == 965  # rounded to a whole number
    assert ws["H5"].value == 965


def test_write_final_output_skips_gh_columns_when_mtd_columns_omitted(template_path, tmp_path):
    values = {"Bed Strength": 1000}
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="F")

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    assert ws["G4"].value is None
    assert ws["H4"].value is None


def test_write_final_output_writes_hospital_revenue_as_a_live_formula(template_path, tmp_path):
    values = {
        "Total Billing": 8742193.15,
        "AEPL Billing": 9667188.25,
        "Hospital Revenue (Net of AEPL)": 8742193.15 - 9667188.25,
    }
    mtd_columns = {
        "Total Billing": {"daily_avg": 485677.4, "mtd_proj": 14570322.0},
        "AEPL Billing": {"daily_avg": 536999.3, "mtd_proj": 16109980.0},
    }
    output_path = tmp_path / "output.xlsx"
    write_final_output(template_path, output_path, values, date_column="F", mtd_columns=mtd_columns)

    wb = openpyxl.load_workbook(output_path)
    ws = wb[SHEET_NAME]
    # row 63 = Hospital Revenue, row 57 = Credit Total Billing, row 60 = AEPL Billing
    assert ws["F63"].value == "=F57-F60"
    assert ws["G63"].value == "=G57-G60"
    assert ws["H63"].value == "=H57-H60"
