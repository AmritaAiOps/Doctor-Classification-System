"""Stage 8: Final output writer.

Writes computed values into the existing 'Final output' sheet at fixed rows,
in a single caller-specified date column. Every row is now resolved -- Long
Stay Patients (row 29) was the last one, computed from the IP Discharges file.
"""
import calendar
import logging
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

logger = logging.getLogger(__name__)

SHEET_NAME = "Final output"

# row number -> key expected in the `values` dict passed to write_final_output
ROW_MAP = {
    4: "Bed Strength",
    5: "Beds Occupied",
    6: "Occupancy %",
    9: "OP New Registration",
    10: "OP Encounters",
    11: "IP Admission Total",
    12: "IP Admission General",
    13: "IP Admission TPA",
    14: "IP Admission ECHS",
    15: "IP Admission P.Card.Fund",
    16: "IP Admission Corporates",
    18: "Emergency Admission",
    19: "Planned Admission",
    20: "Admission from OP (walk-in)",
    22: "IP Discharges Total",
    23: "IP Discharges General",
    24: "IP Discharges TPA",
    25: "IP Discharges ECHS",
    26: "IP Discharges P.Card.Fund",
    27: "IP Discharges Corporates",
    29: "Long Stay Patients",
    32: "OP Billing",
    33: "IP Billing",
    34: "Total Billing",
    37: "Billing Domestic",
    38: "Billing International",
    39: "Billing Total",
    44: "Cash Domestic",
    45: "Cash International",
    48: "Credit Domestic Total",
    49: "Credit Domestic ECHS",
    50: "Credit Domestic P.Card.Fund",
    51: "Credit Domestic TPA",
    52: "Credit Domestic Corporates",
    54: "Credit International Total",
    55: "Credit International TPA Aasantha",
    56: "Credit International Corporates (International)",
    # Row 57 is the grand "Total Billing" at the end of the cash/credit block
    # (= rows 44+45+48+54 = row 34), NOT the credit-only subtotal -- the
    # credit-only figure ("Credit Total Billing") has no row of its own.
    57: "Total Billing",
    60: "AEPL Billing",
    63: "Hospital Revenue (Net of AEPL)",
}

KEY_TO_ROW = {key: row for row, key in ROW_MAP.items()}

# Rows written as a live Excel formula referencing other rows in the SAME
# column, instead of a pre-computed number -- so opening the workbook shows
# staff exactly how the figure was derived. Row 63 = row 57 (grand Total
# Billing) minus row 60 (AEPL Billing).
DERIVED_FORMULA_ROWS = {
    "Hospital Revenue (Net of AEPL)": ("Total Billing", "AEPL Billing"),
}


UNAVAILABLE = "—"  # em dash shown when a month has no recorded days at all


def _copy_column_style(ws, source_col: int, dest_cols) -> None:
    """openpyxl's insert_cols() leaves new cells unstyled -- clone font,
    border, fill, alignment and number format from a neighboring column so
    inserted month columns match the rest of the sheet. Font is forced to
    Arial regardless of the source, per the sheet's house style."""
    for row in range(1, ws.max_row + 1):
        ref = ws.cell(row=row, column=source_col)
        ref_font = ref.font
        for col in dest_cols:
            cell = ws.cell(row=row, column=col)
            cell.font = Font(
                name="Arial", size=ref_font.size, bold=ref_font.bold,
                italic=ref_font.italic, color=ref_font.color,
            )
            cell.border = copy(ref.border)
            cell.fill = copy(ref.fill)
            cell.alignment = copy(ref.alignment)
            cell.number_format = ref.number_format


def _finalize_prior_months(ws, finalized_months: dict) -> int:
    """Inserts a static 2-col ("Daily Average <Month>", "MTD <Month>") pair
    just left of the 'Particulars' column for every completed month not
    already present, filling UNAVAILABLE for months with no recorded days at all.
    Inserting columns never shifts rows, so row numbers stay fixed.
    Returns the (possibly shifted) 'Particulars' column index.
    """
    part_idx = next(
        c for c in range(1, ws.max_column + 1)
        if str(ws.cell(row=1, column=c).value or "").strip() == "Particulars"
    )
    # Only columns LEFT of 'Particulars' are finalized months -- the current
    # 3-col block to the right also has a "Daily Average <Month>" header and
    # must not count as already-finalized.
    existing = {str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, part_idx)}
    # ponytail: headers carry month name only (matches the manual template), so
    # dedupe breaks after 12 months of history -- add the year to headers then.
    for (year, month), finalized in sorted(finalized_months.items()):
        name = calendar.month_name[month]
        if f"Daily Average {name}" in existing:
            continue
        ws.insert_cols(part_idx, 2)
        _copy_column_style(ws, source_col=part_idx + 2, dest_cols=(part_idx, part_idx + 1))
        ws.cell(row=1, column=part_idx).value = f"Daily Average {name}"
        ws.cell(row=1, column=part_idx + 1).value = f"MTD {name}"
        for row, key in ROW_MAP.items():
            if finalized is None:
                ws.cell(row=row, column=part_idx).value = UNAVAILABLE
                ws.cell(row=row, column=part_idx + 1).value = UNAVAILABLE
            elif key in finalized:
                ws.cell(row=row, column=part_idx).value = finalized[key]["daily_avg"]
                ws.cell(row=row, column=part_idx + 1).value = finalized[key]["mtd"]
        part_idx += 2
    return part_idx


def write_final_output(
    template_path,
    output_path,
    values: dict,
    date_column: str = None,
    report_date=None,
    mtd_columns: dict = None,
    finalized_months: dict = None,
) -> str:
    """mtd_columns: optional {row_key: {"daily_avg": ..., "mtd_proj": ...}}
    (see backend.stages.mtd.compute_mtd_columns). When given, writes Daily
    Average into the column right after date_column and MTD (Proj) into the
    one after that -- e.g. date_column="F" writes Daily Average to G and
    MTD (Proj) to H, matching the template's "As on" / "Daily Average" /
    "MTD (Proj)" column layout. mtd_columns=None with finalized_months given
    means the current month has no recorded days at all yet: both columns get "-".

    finalized_months: optional {(year, month): finalize_month(...) result}
    for completed months (see backend.monthly_average). When given, prior
    months roll over into static 2-col pairs left of 'Particulars' and the
    current 3-col block location is derived from the sheet, ignoring
    date_column.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_NAME]

    # The template also carries the raw per-report source sheets (used only
    # as a reference while building the template); the exported workbook
    # should be the Final output sheet alone.
    for sheet_name in list(wb.sheetnames):
        if sheet_name != SHEET_NAME:
            del wb[sheet_name]

    month_gated = finalized_months is not None and mtd_columns is None
    if finalized_months is not None:
        part_idx = _finalize_prior_months(ws, finalized_months)
        date_column = get_column_letter(part_idx + 1)
        if report_date is not None:
            name = calendar.month_name[report_date.month]
            ws.cell(row=1, column=part_idx + 2).value = f"Daily Average {name}"
            ws.cell(row=1, column=part_idx + 3).value = f"MTD {name} (Proj)"

    if report_date is not None:
        ws[f"{date_column}1"] = f"As on {report_date.strftime('%d %b %Y')}"

    date_col_idx = column_index_from_string(date_column)
    daily_avg_column = get_column_letter(date_col_idx + 1)
    mtd_proj_column = get_column_letter(date_col_idx + 2)

    for row, key in ROW_MAP.items():
        if key in DERIVED_FORMULA_ROWS:
            positive_key, negative_key = DERIVED_FORMULA_ROWS[key]
            if positive_key not in KEY_TO_ROW or negative_key not in KEY_TO_ROW:
                logger.warning("write_final_output: cannot build formula for %r; missing row mapping.", key)
                continue
            positive_row, negative_row = KEY_TO_ROW[positive_key], KEY_TO_ROW[negative_key]

            ws[f"{date_column}{row}"] = f"={date_column}{positive_row}-{date_column}{negative_row}"
            if mtd_columns:
                ws[f"{daily_avg_column}{row}"] = f"={daily_avg_column}{positive_row}-{daily_avg_column}{negative_row}"
                ws[f"{mtd_proj_column}{row}"] = f"={mtd_proj_column}{positive_row}-{mtd_proj_column}{negative_row}"
            elif month_gated:
                ws[f"{daily_avg_column}{row}"] = UNAVAILABLE
                ws[f"{mtd_proj_column}{row}"] = UNAVAILABLE
            continue

        if key not in values:
            logger.warning("write_final_output: missing value for %r (row %d); leaving cell untouched.", key, row)
            continue
        ws[f"{date_column}{row}"] = values[key]

        if month_gated:
            ws[f"{daily_avg_column}{row}"] = UNAVAILABLE
            ws[f"{mtd_proj_column}{row}"] = UNAVAILABLE
        elif mtd_columns and key in mtd_columns:
            ws[f"{daily_avg_column}{row}"] = mtd_columns[key]["daily_avg"]
            ws[f"{mtd_proj_column}{row}"] = mtd_columns[key]["mtd_proj"]

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return str(output_path)
